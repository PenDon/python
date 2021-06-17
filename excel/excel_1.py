import time
import random
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import re


# from colorama import init, Fore, Back

def batchExcel(filename):
    """
    Batch processing Excel
    :param filename:
    :return:
    """
    workbook = load_workbook(filename)
    ws = workbook.active
    typeColumn = 0  # 定位产品规格列
    nameColumn = 0  # 定位产品名列
    numberColumn = 0  # 定位产品数量列
    orderColumn = 0  # 定位订单号列
    for row in ws.rows:
        # check column index
        if typeColumn and nameColumn and numberColumn and orderColumn:
            break
        for j, cell in enumerate(row):
            if cell.value == '产品规格':
                typeColumn = j
            if cell.value == '产品名称':
                nameColumn = j
            if cell.value == '产品数量':
                numberColumn = j
            if cell.value == '订单号':
                orderColumn = j
    if not typeColumn or not nameColumn or not numberColumn or not orderColumn:
        print('未检测到列名【产品规格或产品名称或产品数量或订单号】,请检查!')
        # print(color.red('未检测到列名【产品规格或产品名称或产品数量或订单号】,请检查!'))
        time.sleep(5)
        exit(0)
    rows = ws.rows
    for index, row in enumerate(rows):
        if not index:
            continue
        productType = row[typeColumn].value
        productName = row[nameColumn].value

        #  产品名称与产品数量处理
        i = 0  # Gift box & number of 计数
        # print(order.index[0])
        correctName = []  # 存储清除Gift box后的产品名称
        try:
            for name in productName.split('\n'):
                if 'gift box - checked' in name.lower() or name.upper().find('NUMBER OF') != -1:
                    i += 1
                else:
                    if 'variants' in productName.lower():
                        pass
                    else:
                        if 'want dad bead' in name.lower():
                            i += 1
                        else:
                            correctName.append(name)
            row[nameColumn].value = '\n'.join(correctName)
            row[numberColumn].value -= i
        except Exception as e:
            print(f"订单号为【{row[orderColumn].value}】的产品名称无法处理!已跳过")
            # print(color.red(f"订单号为【{row[orderColumn].value}】的产品名称无法处理!已跳过"))
            row[nameColumn].font = font

        #  产品规格处理
        types = []  # 存储处理后的产品规格
        if productType:
            # 无number of的情况，取variants中的数量信息
            if productType.lower().find('number of') == -1:
                productType = re.sub('Variants(.*?)(\d+)( beads| charms| pendants)(.*?)(\n)(.*?)',
                                     r'number of beads:\2\5\6', productType, flags=re.I)
        else:
            continue
        try:
            productType = productType.replace('</n>', '\n')  # 替换</n>
        except Exception as e:
            row[typeColumn].value = ''
        for type in productType.split('\n'):
            if not type:
                continue
            lower = type.lower()

            if type.find('_') == -1 and lower.find('gift box') == -1 and type.find('Variants') == -1:
                flag_special = True
                # 查找 number of 取出具体数量
                if lower.find('number of') != -1:
                    try:
                        type = re.search('\d+', type.split(':')[1]).group() + '颗：'
                    except Exception as e3:
                        print(type)
                    flag_special = False
                    # boy, girl
                if lower.find('charm type') != -1:
                    try:
                        type = ';' + type.split(':')[1]
                        flag_special = False
                    except Exception as e1:
                        print("charm type split error!")
                    # chain length check
                if lower.find('chain length') != -1:
                    try:
                        type = '，链长：' + re.search('\d+CM', type, flags=re.IGNORECASE).group()
                        flag_special = False
                    except Exception as e4:
                        pass
                # birthstone check
                if lower.find('birthstone') != -1 or lower.find('birth month') != -1:
                    type = re.sub(
                        '.*?(birthstone|birth month).*?(January|February|March|April|May|June|July|August|September'
                        '|October|November|December)(.*)',
                        r' + 钻:\2', type, flags=re.I)
                    flag_special = False

                # his name check
                if lower.find('his name') != -1 or lower.split(":")[0] == 'name' or lower.find('her name') != -1:
                    try:
                        type = '大珠右:' + "".join(type.split(':')[-1].split())
                        flag_special = False
                    except Exception as e1:
                        print("his name split error!")

                # custom name check
                if lower.find('name') != -1 and lower.find('his name') == -1 and len(lower.split(":")[0]) > len(
                        'name') and lower.find('father\'s name') == -1 and lower.find('her name') == -1:
                    try:
                        type = '，' + type.split(':')[-1]
                        flag_special = False
                    except Exception as e1:
                        print("name split error!")
                        print(type)

                # father's name check
                if lower.find('father\'s name') != -1:
                    try:
                        type = '大珠:' + "".join(type.split(':')[-1].split())
                        flag_special = False
                    except Exception as e1:
                        print("father\'s name split error!")

                # custom name check
                if lower.find('initial') != -1:
                    try:
                        type = '，' + type.split(':')[1]
                        flag_special = False
                    except Exception as e1:
                        print("initial split error!")
                # custom name check
                if lower.find('inscription') != -1:
                    try:
                        type = '，' + type.split(':')[1]
                        flag_special = False
                    except Exception as e1:
                        print("inscription split error!")
                # size check
                if lower.find('size') != -1 or lower.find('length of') != -1:
                    try:
                        group = re.search('[0-9]+([.][0-9]+)?( )?(mm|cm)', type, re.I)
                        if group:
                            type = '，链长：' + group.group()
                            flag_special = False
                        else:
                            group = re.search('[0-9]+([.][0-9]+)?(\'\'|\")', type)
                            if group:
                                type = '，链长：' + group.group()
                                flag_special = False

                    except Exception as e2:
                        pass
                # check date
                if lower.find('date') != -1:
                    try:
                        type = re.sub('(.*?Date \d+)(.*?)', r' + 日期\2', type)
                        flag_special = False
                    except Exception as e3:
                        pass

                # check 链长
                if lower.find('链长') != -1:
                    flag_special = False

                # check want dad bead
                # show the want dad bead
                if lower.find('want dad bead') != -1:
                    #     type = ''
                    flag_special = False

                # check APPELLATION
                if lower.find('appellation') != -1:
                    try:
                        type = '大珠左:' + "".join(type.split(':')[-1].split())
                        flag_special = False
                    except Exception as e3:
                        print("appellation split error!")
                if index:
                    if flag_special:
                        row[typeColumn].font = font
                        ## add color
                types.append(type)
            else:
                continue
        array = []
        newString = ''
        newTypes = []
        # array = []
        Flag = False
        print(types)
        # 大珠子位置调整到颗数之前
        for index, data in enumerate(types):
            if index == len(types) - 1:
                newTypes.append(data)
                break
            if data.find('颗') != -1:
                if len(data) >= index + 2 and types[index + 1].find('大珠左') != -1 and types[index + 2].find('大珠右') != -1:
                    try:
                        newTypes.extend([types[index + 1], types[index + 2], types[index]])
                        Flag = True
                    except Exception as e5:
                        print("调整大珠显示位置时错误:")
                        print(e5)
                        row[typeColumn].font = font
                else:
                    if len(data) >= index + 1 and types[index + 1].find('大珠') != -1:
                        try:
                            newTypes.extend([types[index + 1], types[index]])
                            Flag = True
                        except Exception as e5:
                            print("调整大珠显示位置时错误:")
                            print(e5)
                            row[typeColumn].font = font
                    else:
                        newTypes.append(data)
            else:
                if Flag and data.find('大珠') != -1:
                    continue
                else:
                    newTypes.append(data)
                    Flag = False

        # 连接名字和颗数到一行
        for type_2 in newTypes:
            if type_2:
                flag_1 = type_2.find('，') == 0
                flag_2 = type_2.find('+') != -1
                flag_3 = type_2.find(';') != -1
                if flag_1 or flag_2 or flag_3:
                    if flag_1 and '大珠' not in type_2:
                        newString += type_2
                    if flag_2:
                        newString += type_2 + '\n'
                    if flag_3:
                        if newString:
                            if newString[-1] != '：':
                                newString += '\n'
                            newString += type_2.replace(';', '')
                        else:
                            print("this:", newTypes)
                else:
                    if newString:
                        newString = newString.replace('：，', '：')
                        array.append(newString)
                    newString = type_2
            else:
                continue
        newString = newString.replace('：，', '：')
        array.append(newString)
        print(array)

        return_string = '\n'.join(array)
        row[typeColumn].value = return_string.replace('\n\n', '\n')
    saveDir = '已处理'
    if os.path.exists(saveDir):
        pass
    else:
        os.mkdir(saveDir)
    excelName = f'./{saveDir}/已处理_' + str(random.randint(1, 9999)) + '_' + filename

    # Excel居中调整/列宽修改，全局居中/自动换行未处理
    ws = workbook[workbook.sheetnames[0]]
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 5
    ws.column_dimensions['I'].width = 5
    ws.column_dimensions['J'].width = 5
    ws.column_dimensions['K'].width = 5
    ws.column_dimensions['L'].width = 5
    ws.column_dimensions['M'].width = 5
    ws.column_dimensions['N'].width = 5
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 10
    ws.column_dimensions['Q'].width = 10
    ws.column_dimensions['R'].width = 10
    workbook.save(excelName)


# color = Colored()
font = Font(color='FF0000')  # Abnormal cells marked in red

allFiles = os.listdir('.')
selection = 9  # selection
i = 1
targetFiles = []
while selection not in range(i):
    i = 1
    targetFiles = []
    print(f'当前目录下存在下列Excel文件等待处理:')
    # print(color.green(f'当前目录下存在下列Excel文件等待处理:'))
    for file in allFiles:
        if file.find('.xlsx') != -1:
            print(f'\t\t{i} : {file}')
            targetFiles.append(file)
            i += 1
    if not targetFiles:
        print('当前目录下未找到需要处理的Excel文件，请检查!')
        # print(color.red('当前目录下未找到需要处理的Excel文件，请检查!'))
        time.sleep(5)
        exit()
    print(f'\t\t0 : 处理所有Excel文件')
    try:
        selection = eval(input("\t" "请选择:"))
        # selection = eval(input("\t" + color.green("请选择:")))
    except Exception as e:
        # print('\t\t' + color.red('请输入一个有效的整数!'))
        print('\t\t' + '请输入一个有效的整数!')
        continue

# if selected(!0)
if selection:
    targetFiles = [targetFiles[selection - 1]]
print(targetFiles)
for file in targetFiles:  # Cyclic treatment
    batchExcel(file)

print('\n全部文件处理完成!')
# print(color.green('全部文件处理完成!'))
time.sleep(5)
